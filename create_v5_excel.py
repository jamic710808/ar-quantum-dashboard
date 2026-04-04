# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ============ 工作表 1: 應收帳款明細 (AR_Detail) ============
ws1 = wb.active
ws1.title = 'AR_Detail'

# 標題列
headers1 = [
    '客戶名稱', '城市', '帳款金額(萬)', '帳齡(天)', '壞帳率(%)',
    '風險等級', '預估損失', '催收策略', '年營收(萬)', '交易次數',
    '信用評分', '業務員', '生命週期', '帳款日期', '到期日'
]

# 樣式
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 寫入標題
for col, header in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 範例資料
sample_data = [
    ['台積電', 'hsinchu', 9.2, 12, 0.08, 'safe', 0, '提升授信', 58.4, 22, 96, '張三', 'mature', '2026-02-15', '2026-03-17'],
    ['聯發科技', 'hsinchu', 5.8, 18, 0.3, 'low', 29, '維持現狀', 35.6, 15, 82, '李四', 'mature', '2026-02-20', '2026-03-20'],
    ['鴻海精密', 'taipei', 3.4, 25, 0.4, 'low', 17, '維持現狀', 22.8, 18, 88, '王五', 'mature', '2026-02-10', '2026-03-12'],
    ['日月光半導體', 'kaohsiung', 6.3, 22, 0.6, 'mid', 94, '加強催收', 28.3, 9, 72, '趙六', 'growing', '2026-02-25', '2026-03-25'],
    ['宏碁資訊服務', 'taipei', 4.1, 28, 0.8, 'mid', 82, '加強催收', 18.5, 6, 65, '王五', 'growing', '2026-02-08', '2026-03-10'],
    ['廣達電腦', 'taichung', 2.1, 32, 0.9, 'mid', 42, '加強催收', 14.2, 7, 61, '趙六', 'growing', '2026-02-05', '2026-03-07'],
    ['環球晶圓(股)', 'hsinchu', 8.2, 38, 1.5, 'high', 246, '立即清收', 24.1, 12, 42, '李四', 'declining', '2026-01-28', '2026-03-07'],
    ['三星電子-行動部門', 'kaohsiung', 12.5, 45, 2.5, 'critical', 625, '現款結算', 38.2, 8, 28, '張三', 'declining', '2026-01-20', '2026-03-06'],
    ['華碩電腦', 'taipei', 2.8, 8, 0.1, 'safe', 0, '提升授信', 12.3, 3, 78, '陳七', 'new', '2026-03-01', '2026-03-31'],
    ['仁寶電腦', 'taichung', 1.8, 15, 0.2, 'safe', 0, '提升授信', 10.4, 11, 85, '陳七', 'mature', '2026-02-28', '2026-03-15'],
]

# 寫入範例資料
for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')

# 設定欄寬
col_widths1 = [20, 12, 14, 10, 10, 12, 12, 14, 14, 12, 10, 10, 12, 12, 12]
for i, width in enumerate(col_widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = width

# 資料驗證 - 風險等級下拉
dv_risk = DataValidation(type='list', formula1='"safe,low,mid,high,critical"', allow_blank=True)
ws1.add_data_validation(dv_risk)
dv_risk.add('F2:F1000')

# 資料驗證 - 生命週期下拉
dv_lifecycle = DataValidation(type='list', formula1='"new,growing,mature,declining"', allow_blank=True)
ws1.add_data_validation(dv_lifecycle)
dv_lifecycle.add('M2:M1000')

# ============ 工作表 2: 客戶信用歷史 (Credit_History) ============
ws2 = wb.create_sheet('Credit_History')

headers2 = ['客戶名稱', '月份', '信用評分', '帳款金額(萬)', '帳齡(天)', '風險等級', '生命週期']

for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 範例歷史資料 (12個月)
months = ['2025-03', '2025-04', '2025-05', '2025-06', '2025-07', '2025-08',
          '2025-09', '2025-10', '2025-11', '2025-12', '2026-01', '2026-02']

history_data = []
for customer in ['台積電', '聯發科技', '鴻海精密', '日月光半導體', '宏碁資訊服務']:
    base_score = 90 if customer == '台積電' else 75
    for i, month in enumerate(months):
        score = base_score - (i % 3) * 5 + (i % 2) * 3
        if score > 100: score = 100
        if score < 30: score = 30
        history_data.append([customer, month, score, 8 - i*0.3, 15 + i*2,
                           'safe' if score > 80 else 'mid' if score > 50 else 'high',
                           'mature' if i > 6 else 'growing'])

for row_idx, row_data in enumerate(history_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

for i, width in enumerate([20, 10, 10, 14, 10, 10, 12], 1):
    ws2.column_dimensions[get_column_letter(i)].width = width

# ============ 工作表 3: 業務員資料 (Salesperson) ============
ws3 = wb.create_sheet('Salesperson')

headers3 = ['業務員', '部門', '入職日期', '負責客戶數', '總帳款(萬)', '平均帳齡', '壞帳率(%)', '績效評分']

for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

salesperson_data = [
    ['張三', '業務一部', '2022-03-15', 2, 21.7, 28.5, 1.29, 85],
    ['李四', '業務二部', '2021-08-20', 2, 14.0, 28.0, 0.90, 78],
    ['王五', '業務一部', '2023-01-10', 2, 7.5, 26.5, 0.60, 82],
    ['趙六', '業務三部', '2022-06-01', 2, 8.4, 27.0, 0.75, 80],
    ['陳七', '業務二部', '2023-09-05', 2, 4.6, 11.5, 0.15, 92],
]

for row_idx, row_data in enumerate(salesperson_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

for i, width in enumerate([12, 12, 12, 12, 12, 12, 12, 12], 1):
    ws3.column_dimensions[get_column_letter(i)].width = width

# ============ 工作表 4: 系統參數 (Settings) ============
ws4 = wb.create_sheet('Settings')

settings_data = [
    ['參數名稱', '數值', '說明'],
    ['帳齡閾值1', 30, '第一段階帳齡天數'],
    ['帳齡閾值2', 60, '第二段階帳齡天數'],
    ['帳齡閾值3', 90, '第三段階帳齡天數'],
    ['帳齡閾值4', 120, '第四段階帳齡天數'],
    ['', '', ''],
    ['信用評分權重_帳齡', 0.3, '帳齡因素權重'],
    ['信用評分權重_金額', 0.25, '金額因素權重'],
    ['信用評分權重_交易次數', 0.2, '交易次數權重'],
    ['信用評分權重_風險', 0.25, '風險等級權重'],
    ['', '', ''],
    ['風險閾值_safe', 80, '安全客戶評分門檻'],
    ['風險閾值_low', 60, '低風險客戶評分門檻'],
    ['風險閾值_mid', 40, '中等風險客戶評分門檻'],
    ['風險閾值_high', 20, '高風險客戶評分門檻'],
    ['', '', ''],
    ['預測_歷史月數', 12, '用於預測的歷史月數'],
    ['預測_信心度', 0.95, '預測區間信心度'],
    ['WhatIf_利率範圍_min', 0, '利率滑桿最小值(%)'],
    ['WhatIf_利率範圍_max', 15, '利率滑桿最大值(%)'],
    ['WhatIf_催收強度_min', 0.5, '催收強度最小值(x)'],
    ['WhatIf_催收強度_max', 2.0, '催收強度最大值(x)'],
]

for row_idx, row_data in enumerate(settings_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.alignment = Alignment(horizontal='left')

ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 12
ws4.column_dimensions['C'].width = 30

# 隱藏網格線 (除了 AR_Detail)
ws2.sheet_view.showGridLines = False
ws3.sheet_view.showGridLines = False
ws4.sheet_view.showGridLines = False

# 儲存檔案
wb.save('AR_Template_V5_0.xlsx')
print('[OK] AR_Template_V5_0.xlsx created successfully!')
